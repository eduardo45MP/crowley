from __future__ import annotations

import csv
import io
import json
import zipfile
from pathlib import Path
from typing import Any

from market_intelligence.editorial.models import PublishedOpportunity, PublishedReport


class ReportExporter:
    def export(self, report: PublishedReport, output_root: str | Path, formats: list[str]) -> dict[str, Path]:
        requested = [item.strip().lower() for item in formats if item.strip()]
        unsupported = sorted(set(requested) - {"json", "csv", "xlsx", "pdf"})
        if unsupported:
            raise ValueError(f"Formatos não suportados: {', '.join(unsupported)}")
        report_dir = Path(output_root) / report.snapshot.report_id
        if report_dir.exists() and any(report_dir.iterdir()):
            raise FileExistsError(f"O snapshot {report.snapshot.report_id} já existe em {report_dir}; snapshots publicados são imutáveis.")
        report_dir.mkdir(parents=True, exist_ok=True)
        paths: dict[str, Path] = {}
        if "json" in requested:
            paths["json"] = self.export_json(report, report_dir / "report.json")
        if "csv" in requested:
            paths["csv"] = self.export_csv(report, report_dir / "opportunities.csv")
        if "xlsx" in requested:
            paths["xlsx"] = self.export_xlsx(report, report_dir / "crowley-opportunities.xlsx")
        if "pdf" in requested:
            paths["pdf"] = self.export_pdf(report, report_dir / "crowley-report.pdf")
        return paths

    @staticmethod
    def export_json(report: PublishedReport, path: Path) -> Path:
        text = json.dumps(report.as_dict(), ensure_ascii=False, indent=2, sort_keys=True, allow_nan=False) + "\n"
        path.write_text(text, encoding="utf-8", newline="\n")
        return path

    @staticmethod
    def export_csv(report: PublishedReport, path: Path) -> Path:
        headers = [
            "Rank", "Product", "Niche", "Buyer Group", "Product Type", "Demand Score",
            "Competition Score", "Purchase Intent", "Build Ease", "Differentiation",
            "Opportunity Score", "Opportunity Confidence", "Research Confidence",
            "Observed Median Price", "Recommended Price", "Build Hours",
            "Revenue Efficiency Score", "Keywords",
        ]
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle, lineterminator="\n")
            writer.writerow(headers)
            for item in report.ranking:
                writer.writerow([
                    item.rank, item.product_name, item.niche, item.buyer_group, item.product_type,
                    _csv_number(item.demand_score), _csv_number(item.competition_score),
                    _csv_number(item.purchase_intent_score), _csv_number(item.build_ease_score),
                    _csv_number(item.differentiation_score), _csv_number(item.opportunity_score),
                    _csv_number(item.opportunity_confidence), _csv_number(item.research_confidence),
                    _csv_number(item.price_median), _csv_number(item.recommended_price),
                    _csv_number(item.estimated_build_hours), _csv_number(item.revenue_efficiency_score),
                    " | ".join(item.keywords),
                ])
        return path

    @staticmethod
    def export_xlsx(report: PublishedReport, path: Path) -> Path:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise RuntimeError("XLSX requer openpyxl. Instale as dependências com `pip install -r requirements.txt`.") from exc

        workbook = Workbook()
        top100 = workbook.active
        top100.title = "Top 100"
        top10 = workbook.create_sheet("Top 10")
        methodology = workbook.create_sheet("Methodology")
        metadata = workbook.create_sheet("Metadata")
        header_fill = PatternFill("solid", fgColor="17324D")
        header_font = Font(color="FFFFFF", bold=True)
        accent_fill = PatternFill("solid", fgColor="DCEAF7")

        ranking_headers = [
            "Rank", "Product", "Niche", "Buyer Group", "Product Type", "Problem", "Demand",
            "Competition", "Purchase Intent", "Build Ease", "Differentiation", "Opportunity Score",
            "Opportunity Confidence", "Research Confidence", "Median Price", "Recommended Price",
            "Build Hours", "Build Complexity", "Revenue Efficiency", "Keywords",
        ]
        top100.append(ranking_headers)
        for item in report.ranking:
            top100.append([
                item.rank, item.product_name, item.niche, item.buyer_group, item.product_type,
                item.problem, item.demand_score, item.competition_score, item.purchase_intent_score,
                item.build_ease_score, item.differentiation_score, item.opportunity_score,
                item.opportunity_confidence, item.research_confidence, item.price_median,
                item.recommended_price, item.estimated_build_hours, item.build_complexity,
                item.revenue_efficiency_score, " | ".join(item.keywords),
            ])

        top10_headers = [
            "Top 10 Rank", "Selection Rank", "Product", "Opportunity Score", "Opportunity Thesis",
            "Buyer / Problem", "Positioning", "Pricing", "Keywords", "Build Hours",
            "Differentiators", "Warnings", "Blueprint Summary",
        ]
        top10.append(top10_headers)
        for item in report.top10:
            thesis = (item.thesis or {}).get("opportunity_statement")
            pricing = _pricing_text(item)
            blueprint = item.blueprint or {}
            summary = "; ".join(blueprint.get("mvp_features") or [])
            top10.append([
                item.top10_rank, item.selection_rank, item.product_name, item.opportunity_score, thesis,
                " / ".join(value for value in [item.target_buyer, item.problem] if value), item.positioning,
                pricing, " | ".join(item.keywords), item.estimated_build_hours, item.differentiation,
                " | ".join(item.warnings), summary,
            ])

        methodology.append(["Crowley Methodology", report.methodology.get("version")])
        methodology.append(["Stage", "Definition"])
        for name, definition in report.methodology.get("stages", {}).items():
            methodology.append([name.replace("_", " ").title(), definition])
        methodology.append([])
        methodology.append(["Limitations"])
        for limitation in report.methodology.get("limitations", []):
            methodology.append([limitation])

        metadata.append(["Field", "Value"])
        for key, value in report.snapshot.as_dict().items():
            metadata.append([key, json.dumps(value, ensure_ascii=False, sort_keys=True) if isinstance(value, (dict, list)) else value])

        for sheet in workbook.worksheets:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions if sheet.max_row > 1 and sheet.max_column > 1 else None
            sheet.sheet_view.showGridLines = False
            for cell in sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            if sheet.title == "Methodology":
                for cell in sheet[2]:
                    cell.fill = accent_fill
                    cell.font = Font(bold=True)
            for row in sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
            for column_index, cells in enumerate(sheet.columns, start=1):
                max_length = min(55, max(10, max(len(str(cell.value or "")) for cell in cells) + 2))
                sheet.column_dimensions[get_column_letter(column_index)].width = max_length
            sheet.row_dimensions[1].height = 30
            sheet.sheet_properties.pageSetUpPr.fitToPage = True
            sheet.page_setup.fitToWidth = 1
            sheet.page_setup.fitToHeight = 0
            sheet.page_setup.orientation = "landscape" if sheet.title in {"Top 100", "Top 10"} else "portrait"
            sheet.print_title_rows = "1:1"
            sheet.print_area = sheet.dimensions

        for index, width in enumerate([7, 28, 16, 18, 14, 22, 10, 11, 12, 10, 13, 14, 14, 14, 13, 14, 11, 14, 14, 36], start=1):
            top100.column_dimensions[get_column_letter(index)].width = width
        for index, width in enumerate([10, 10, 28, 14, 45, 32, 45, 34, 35, 12, 28, 34, 40], start=1):
            top10.column_dimensions[get_column_letter(index)].width = width
        methodology.column_dimensions["A"].width = 26
        methodology.column_dimensions["B"].width = 100
        metadata.column_dimensions["A"].width = 32
        metadata.column_dimensions["B"].width = 90

        for sheet in (top100, top10):
            for row in range(2, sheet.max_row + 1):
                for column in range(7 if sheet is top100 else 4, min(sheet.max_column, 19) + 1):
                    cell = sheet.cell(row=row, column=column)
                    if isinstance(cell.value, float):
                        cell.number_format = "0.00"
        for row in range(2, top100.max_row + 1):
            top100.cell(row, 13).number_format = "0.0%"
            top100.cell(row, 14).number_format = "0.0%"
            top100.cell(row, 15).number_format = '"$"#,##0.00'
            top100.cell(row, 16).number_format = '"$"#,##0.00'

        workbook.properties.creator = "Crowley"
        workbook.properties.created = report.snapshot.created_at.replace(tzinfo=None)
        workbook.properties.modified = report.snapshot.created_at.replace(tzinfo=None)
        workbook.save(path)
        _normalize_xlsx(path)
        return path

    @staticmethod
    def export_pdf(report: PublishedReport, path: Path) -> Path:
        try:
            from reportlab import rl_config
            rl_config.invariant = 1
            from reportlab.lib import colors
            from reportlab.lib.enums import TA_CENTER
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import mm
            from reportlab.platypus import LongTable, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
        except ImportError as exc:
            raise RuntimeError("PDF requer reportlab. Instale as dependências com `pip install -r requirements.txt`.") from exc

        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name="CoverTitle", parent=styles["Title"], alignment=TA_CENTER, fontSize=24, leading=29, spaceAfter=20))
        styles.add(ParagraphStyle(name="Small", parent=styles["BodyText"], fontSize=8, leading=10))
        styles["BodyText"].leading = 14
        document = SimpleDocTemplate(
            str(path), pagesize=A4, rightMargin=18 * mm, leftMargin=18 * mm,
            topMargin=18 * mm, bottomMargin=18 * mm,
            title="Crowley Opportunity Report", author="Crowley", invariant=1,
        )
        story: list[Any] = []
        story.extend([
            Spacer(1, 45 * mm), Paragraph("Crowley Opportunity Report", styles["CoverTitle"]),
            Paragraph(f"Editorial snapshot {report.snapshot.report_id}", styles["Heading2"]),
            Spacer(1, 10 * mm),
            Paragraph(f"{len(report.ranking)} evidence-backed opportunities; {len(report.top10)} detailed Top 10 dossiers.", styles["BodyText"]),
            PageBreak(),
            Paragraph("Executive Summary", styles["Heading1"]),
            Paragraph(f"This report publishes {len(report.ranking)} opportunities in the exact order produced by the diversity-aware Selection stage. Deep Research enriches the Top 10 without changing upstream scores.", styles["BodyText"]),
            Paragraph("How to Use This Report", styles["Heading1"]),
            Paragraph("Use Opportunity Score for comparative attractiveness, confidence and evidence fields for uncertainty, build hours for scope, and the source references for audit. Treat pricing and Revenue Efficiency as editorial heuristics, not forecasts.", styles["BodyText"]),
            Paragraph("Methodology", styles["Heading1"]),
        ])
        for name, definition in report.methodology.get("stages", {}).items():
            story.append(Paragraph(f"<b>{_escape(name.replace('_', ' ').title())}:</b> {_escape(definition)}", styles["BodyText"]))
        story.append(PageBreak())
        story.append(Paragraph("Top 10 Opportunities", styles["Heading1"]))
        for item in report.top10:
            story.extend(_top10_story(item, styles, Paragraph, Spacer, Table, TableStyle, colors, mm))
            story.append(PageBreak())

        story.append(Paragraph("Ranking 11-100", styles["Heading1"]))
        remaining = [item for item in report.ranking if item.top10_rank is None]
        table_data = [["Rank", "Product", "Niche", "Score", "Confidence", "Build h", "Efficiency"]]
        for item in remaining:
            table_data.append([
                item.rank, Paragraph(_escape(item.product_name), styles["Small"]), _safe(item.niche),
                _format(item.opportunity_score), _format(item.opportunity_confidence),
                _format(item.estimated_build_hours), _format(item.revenue_efficiency_score),
            ])
        ranking_table = LongTable(table_data, repeatRows=1, colWidths=[12 * mm, 58 * mm, 32 * mm, 18 * mm, 20 * mm, 17 * mm, 20 * mm])
        ranking_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17324D")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 7),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F6F8")]),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.HexColor("#17324D")),
        ]))
        story.append(ranking_table)
        story.extend([
            PageBreak(), Paragraph("Methodology Notes", styles["Heading1"]),
            Paragraph("All editorial fields are derived from persisted upstream artifacts. Missing evidence remains blank; the reporting layer does not recalculate market-intelligence dimensions.", styles["BodyText"]),
            Paragraph("Limitations", styles["Heading1"]),
        ])
        for limitation in report.methodology.get("limitations", []):
            story.append(Paragraph(f"- {_escape(limitation)}", styles["BodyText"]))
        story.extend([
            Paragraph("Provenance / Report Metadata", styles["Heading1"]),
            Paragraph(_escape(json.dumps(report.snapshot.as_dict(), ensure_ascii=False, sort_keys=True)), styles["Small"]),
            Paragraph(_escape(report.provenance.get("trace")), styles["Small"]),
        ])

        def footer(canvas: Any, doc: Any) -> None:
            canvas.saveState()
            canvas.setFont("Helvetica", 8)
            canvas.drawString(18 * mm, 9 * mm, report.snapshot.report_id)
            canvas.drawRightString(A4[0] - 18 * mm, 9 * mm, f"Page {doc.page}")
            canvas.restoreState()

        document.build(story, onFirstPage=footer, onLaterPages=footer)
        return path


def _top10_story(item: PublishedOpportunity, styles: Any, Paragraph: Any, Spacer: Any, Table: Any, TableStyle: Any, colors: Any, mm: Any) -> list[Any]:
    thesis = (item.thesis or {}).get("opportunity_statement") or "No persisted Opportunity Thesis available."
    blueprint = item.blueprint or {}
    cell = lambda value: Paragraph(_escape(value), styles["Small"])
    label = lambda value: Paragraph(f"<b>{_escape(value)}</b>", styles["Small"])
    rows = [
        [label("Opportunity Score"), cell(_format(item.opportunity_score)), label("Revenue Efficiency"), cell(_format(item.revenue_efficiency_score))],
        [label("Buyer"), cell(item.target_buyer), label("Problem"), cell(item.problem)],
        [label("Pricing"), cell(_pricing_text(item)), label("Build effort"), cell(f"{_format(item.estimated_build_hours)} h / {_safe(item.build_complexity)}")],
        [label("Keywords"), cell(", ".join(item.keywords)), label("Warnings"), cell("; ".join(item.warnings) or "None recorded")],
    ]
    table = Table(rows, colWidths=[24 * mm, 61 * mm, 24 * mm, 63 * mm])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#DCEAF7")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#DCEAF7")),
        ("FONTSIZE", (0, 0), (-1, -1), 8), ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#AAB7C4")),
    ]))
    return [
        Paragraph(f"{item.top10_rank}. {_escape(item.product_name)}", styles["Heading2"]),
        Paragraph(f"<b>Opportunity Thesis:</b> {_escape(thesis)}", styles["BodyText"]),
        Paragraph(f"<b>Positioning:</b> {_escape(item.positioning)}", styles["BodyText"]),
        table, Spacer(1, 4 * mm),
        Paragraph(f"<b>Evidence summary:</b> {_escape('; '.join(item.evidence_refs[:8]))}", styles["Small"]),
        Paragraph(f"<b>Blueprint summary:</b> {_escape('; '.join(blueprint.get('mvp_features') or []))}", styles["BodyText"]),
    ]


def _normalize_xlsx(path: Path) -> None:
    buffer = io.BytesIO()
    with zipfile.ZipFile(path, "r") as source, zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as target:
        for name in sorted(source.namelist()):
            info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.external_attr = source.getinfo(name).external_attr
            target.writestr(info, source.read(name))
    path.write_bytes(buffer.getvalue())


def _csv_number(value: float | None) -> str:
    return "" if value is None else f"{float(value):.4f}".rstrip("0").rstrip(".")


def _format(value: float | None) -> str:
    return "n/a" if value is None else f"{float(value):.2f}"


def _safe(value: Any) -> str:
    return "n/a" if value is None or str(value).strip() == "" else str(value)


def _escape(value: Any) -> str:
    from xml.sax.saxutils import escape
    return escape(_safe(value))


def _pricing_text(item: PublishedOpportunity) -> str:
    if item.price_median is None:
        return "No observed pricing evidence"
    currency = item.price_currency or ""
    return f"Observed {currency} {_format(item.price_min)} / {_format(item.price_median)} / {_format(item.price_max)}; recommended {_format(item.recommended_price)}"
