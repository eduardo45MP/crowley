from __future__ import annotations

import csv
import json
from datetime import datetime, timezone
from decimal import Decimal

from openpyxl import load_workbook

from crawler.clustering import ProductCluster, ProductClusterMembership
from crawler.models import Marketplace, Product, RawMarketplaceProduct
from crawler.repositories.sqlalchemy_repository import SqlAlchemyProductRepository
from market_intelligence.build_ease.service import BuildEaseAnalysisService
from market_intelligence.competition.service import CompetitionAnalysisService
from market_intelligence.deep_research.service import DeepResearchService
from market_intelligence.demand.service import DemandScoringService
from market_intelligence.differentiation.service import DifferentiationAnalysisService
from market_intelligence.editorial.metrics import revenue_efficiency_score, summarize_pricing
from market_intelligence.editorial.positioning import derive_commercial_positioning
from market_intelligence.editorial.service import EditorialReportService, editorial_keywords
from market_intelligence.opportunity.service import OpportunityAnalysisService
from market_intelligence.product_blueprint.service import ProductBlueprintService
from market_intelligence.purchase_intent.service import PurchaseIntentAnalysisService
from market_intelligence.reporting.exporters import ReportExporter
from market_intelligence.selection.config import SelectionPolicy
from market_intelligence.selection.models import OpportunityCandidate
from market_intelligence.selection.service import PortfolioSelector
from market_intelligence.top10.config import Top10SelectionConfig
from market_intelligence.top10.selection import Top10Selector
from market_intelligence.top10.thesis import OpportunityThesisService


FIXED_TIME = datetime(2026, 8, 26, 12, 0, tzinfo=timezone.utc)


def test_pricing_summary_preserves_observations_and_bounds_recommendation():
    result = summarize_pricing({"minimum": 10, "median": 20, "maximum": 21, "currency": "USD"})
    assert result.minimum_observed_price == 10
    assert result.median_observed_price == 20
    assert result.maximum_observed_price == 21
    assert result.recommended_price == 21
    assert summarize_pricing({}).recommended_price is None


def test_revenue_efficiency_is_bounded_monotonic_and_zero_safe():
    baseline = revenue_efficiency_score(80, 10)
    assert baseline is not None and 0 <= baseline <= 100
    assert revenue_efficiency_score(90, 10) > baseline
    assert revenue_efficiency_score(80, 20) < baseline
    assert revenue_efficiency_score(80, 0) is not None
    assert revenue_efficiency_score(None, 10) is None


def test_positioning_and_keywords_use_only_supplied_artifacts():
    cluster = type("Cluster", (), {"name": "Bakery Pricing", "niche": "bakery", "primary_problem": "pricing", "keywords": ["Bakery", "pricing"]})()
    thesis = {"target_buyer": "home bakers", "problem": "manual pricing", "proposed_advantage": ["labor costing"]}
    blueprint = {"product_name": "Bakery Margin Tool", "value_proposition": "clarify cost and margin", "differentiation_features": ["waste tracking"]}
    dossier = type("Dossier", (), {"differentiation_axes": ["workflow"], "keyword_analysis": {"top_keywords": ["Pricing", "margin"], "long_tail_keywords": []}, "competitor_profiles": []})()
    result = derive_commercial_positioning(cluster=cluster, buyer_group="small_business", thesis=thesis, blueprint=blueprint, dossier=dossier)
    assert result.suggested_product_name == "Bakery Margin Tool"
    assert result.target_buyer == "home bakers"
    assert result.primary_differentiator == "waste tracking"
    assert "revenue" not in (result.short_positioning_statement or "").lower()
    assert editorial_keywords(cluster, dossier) == ["Pricing", "margin", "Bakery"]


def test_end_to_end_persisted_editorial_report_exports(tmp_path):
    repository = SqlAlchemyProductRepository(f"sqlite+pysqlite:///{tmp_path / 'crowley.db'}")
    repository.create_schema()
    clusters = [_persist_cluster(repository, index) for index in range(1, 4)]

    DemandScoringService(repository=repository).calculate(clusters)
    CompetitionAnalysisService(repository=repository).analyze(clusters)
    PurchaseIntentAnalysisService(repository=repository).analyze(clusters)
    BuildEaseAnalysisService(repository=repository).analyze(clusters)
    DifferentiationAnalysisService(repository=repository).analyze(clusters)
    opportunity_result = OpportunityAnalysisService(repository=repository).analyze(clusters)

    candidates = []
    for cluster in clusters:
        persisted = repository.latest_cluster_opportunity_score(cluster.id)
        candidates.append(OpportunityCandidate(
            cluster_id=cluster.id,
            cluster_name=cluster.name,
            opportunity_score=persisted["opportunity_score"],
            opportunity_confidence=persisted["opportunity_confidence"],
            evidence_coverage=persisted["evidence_coverage"],
            buyer_group="small_business" if cluster.id % 2 else "creators",
            niche=cluster.niche,
            problem_type=cluster.primary_problem,
            product_type=cluster.product_type,
            ranking_eligible=True,
        ))
    selection = PortfolioSelector(policy=SelectionPolicy(
        target_size=3,
        minimum_opportunity_score=0,
        minimum_confidence=0,
        minimum_evidence_coverage=0,
        buyer_group_quotas={},
    )).select(candidates)
    repository.save_selection_run(selection.run)
    for item in selection.selected:
        repository.save_selected_opportunity(item, selection.run.id)

    selected_clusters = [repository.get_cluster_by_id(item.cluster_id) for item in selection.selected]
    research = DeepResearchService().run(selected_clusters, top=3, selection_run_id=selection.run.id)
    repository.save_deep_research_run(research.run)
    for dossier in research.dossiers:
        dossier.run_id = research.run.id
        repository.save_deep_research_dossier(dossier)

    opportunity_scores = {cluster.id: repository.latest_cluster_opportunity_score(cluster.id)["opportunity_score"] for cluster in clusters}
    build_scores = {cluster.id: repository.latest_cluster_build_ease_score(cluster.id)["build_ease_score"] for cluster in clusters}
    top10 = Top10Selector(Top10SelectionConfig(target_size=2, minimum_research_coverage=0, minimum_research_confidence=0)).select(
        research.dossiers, opportunity_scores=opportunity_scores, build_ease_scores=build_scores,
    )
    top10.run.deep_research_run_id = research.run.id
    repository.save_top10_run(top10.run)
    for item in top10.selected:
        repository.save_top10_opportunity(item, top10.run.id)
        cluster = repository.get_cluster_by_id(item.cluster_id)
        dossier = next(value for value in research.dossiers if value.cluster_id == item.cluster_id)
        thesis = OpportunityThesisService().create(cluster, dossier, item.opportunity_score)
        repository.save_thesis(thesis)
        repository.save_blueprint(ProductBlueprintService().create(cluster, dossier, thesis))

    score_before = repository.latest_cluster_opportunity_score(clusters[0].id)["opportunity_score"]
    service = EditorialReportService(repository)
    report = service.build(selection_run_id=selection.run.id, top=100, top10_count=2, created_at=FIXED_TIME)
    repeated = service.build(selection_run_id=selection.run.id, top=100, top10_count=2, created_at=FIXED_TIME)
    assert report.as_dict() == repeated.as_dict()
    assert [item.cluster_id for item in report.ranking] == [item.cluster_id for item in selection.selected]
    assert [item.cluster_id for item in report.top10] == [item.cluster_id for item in top10.selected]
    assert report.snapshot.opportunity_count == 3
    assert report.as_dict()["summary"]["fewer_than_requested"] is True
    assert repository.latest_cluster_opportunity_score(clusters[0].id)["opportunity_score"] == score_before
    assert all(item.price_median is not None for item in report.ranking), [(item.cluster_id, item.price_median, item.evidence_refs) for item in report.ranking]
    assert all(item.estimated_build_hours is None for item in report.ranking if item.top10_rank is None)

    paths = ReportExporter().export(report, tmp_path / "reports", ["json", "csv", "xlsx", "pdf"])
    assert set(paths) == {"json", "csv", "xlsx", "pdf"}
    assert all(path.exists() and path.stat().st_size > 0 for path in paths.values())
    payload = json.loads(paths["json"].read_text(encoding="utf-8"))
    assert [row["cluster_id"] for row in payload["ranking"]] == [item.cluster_id for item in selection.selected]
    with paths["csv"].open(encoding="utf-8", newline="") as handle:
        assert len(list(csv.DictReader(handle))) == 3
    workbook = load_workbook(paths["xlsx"], data_only=False)
    assert workbook.sheetnames == ["Top 100", "Top 10", "Methodology", "Metadata"]
    assert workbook["Top 100"].max_row == 4
    assert paths["pdf"].read_bytes().startswith(b"%PDF")
    repeated_json = ReportExporter.export_json(report, tmp_path / "repeated.json")
    repeated_csv = ReportExporter.export_csv(report, tmp_path / "repeated.csv")
    repeated_xlsx = ReportExporter.export_xlsx(report, tmp_path / "repeated.xlsx")
    repeated_pdf = ReportExporter.export_pdf(report, tmp_path / "repeated.pdf")
    assert repeated_json.read_bytes() == paths["json"].read_bytes()
    assert repeated_csv.read_bytes() == paths["csv"].read_bytes()
    assert repeated_xlsx.read_bytes() == paths["xlsx"].read_bytes()
    assert repeated_pdf.read_bytes() == paths["pdf"].read_bytes()


def _persist_cluster(repository: SqlAlchemyProductRepository, index: int) -> ProductCluster:
    members = []
    for member_index, price in enumerate(("19.00", "29.00", "49.00"), start=1):
        external_id = f"cluster-{index}-product-{member_index}"
        raw = repository.save_raw(RawMarketplaceProduct(
            marketplace=Marketplace.MOCK,
            external_id=external_id,
            query="bakery pricing calculator",
            raw_payload={"name": external_id, "price": price},
            collected_at=FIXED_TIME,
        ))
        product = Product(
            product_name=f"Bakery Pricing Tool {index}-{member_index}",
            marketplace=Marketplace.MOCK,
            url=f"https://example.test/{external_id}",
            collected_at=FIXED_TIME,
            external_id=external_id,
            niche=f"bakery-{index}",
            product_type=None,
            price=Decimal(price),
            currency="USD",
            review_count=100 * member_index,
            rating=4.5 + member_index / 10,
            seller=f"seller-{member_index}",
            keywords=["bakery", "pricing", "calculator", f"segment-{index}"],
            description="Pricing calculator with labor, cost, margin, waste, dashboard and editable workflow.",
            image_urls=[f"https://example.test/{external_id}.png"],
            raw_product_id=raw.id,
        )
        saved = repository.upsert_product(product).product
        members.append(saved)
    cluster = repository.save_cluster(ProductCluster(
        name=f"Bakery Pricing Calculator {index}",
        slug=f"bakery-pricing-calculator-{index}",
        niche=f"bakery-{index}",
        product_type="calculator",
        primary_problem="pricing",
        keywords=["bakery", "pricing", "calculator", f"segment-{index}"],
        product_count=len(members),
        confidence=0.9,
        members=members,
    ))
    for product in members:
        repository.save_membership(ProductClusterMembership(product=product, cluster_id=cluster.id, membership_score=0.9, created_at=FIXED_TIME))
    return repository.get_cluster_by_id(cluster.id)
